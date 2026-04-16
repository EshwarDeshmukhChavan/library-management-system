import openpyxl
import sys
import traceback

try:
    print("Loading workbook...")
    sys.stdout.flush()
    wb = openpyxl.load_workbook(
        r'd:\library management system\technical coding 2Library Management 1.xlsx',
        data_only=True,
        read_only=True
    )
    print(f"Sheet names: {wb.sheetnames}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    sys.stdout.flush()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*80}")
        sys.stdout.flush()

        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_data = [str(c) if c is not None else "" for c in row]
            if any(r.strip() for r in row_data):
                print(" | ".join(row_data))
                row_count += 1
        print(f"--- Total non-empty rows: {row_count} ---")
        sys.stdout.flush()

    wb.close()
    print("\nDone!")

except Exception as e:
    traceback.print_exc()
    sys.exit(1)
